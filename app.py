import re
import os
import json
import uuid
import time
import random
from io import BytesIO
from datetime import datetime, timezone
from urllib.parse import urljoin
from typing import Optional, List, Dict, Any

import pandas as pd
import requests
from bs4 import BeautifulSoup
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# -------------------------------------------------------------------
# Constants / Layout knobs
# -------------------------------------------------------------------

st.set_page_config(page_title="Amazon Top-20 Scraper", layout="wide")

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
]

SAVED_DIR = ".saved_searches"
os.makedirs(SAVED_DIR, exist_ok=True)

# Excel layout knobs
EXCEL_COL_WIDTH = 24
EXCEL_IMG_MAX_PX = 220
EXCEL_IMG_ROW_HEIGHT = 90
BAND_COLOR_1 = "FFFFFF"
BAND_COLOR_2 = "F7F7F7"
RANK_BG = "C45500"  # Amazon orange for rank rows
LABEL_BOLD = True

# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------

def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"Accept-Language": "en-US,en;q=0.9"})
    return s


def utc_now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def get_soup(url: str, session: Optional[requests.Session] = None, timeout: int = 15):
    session = session or _session()
    headers = {"User-Agent": random.choice(USER_AGENTS)}
    r = session.get(url, headers=headers, timeout=timeout)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser"), r.url


# -------------------------------------------------------------------
# Amazon scraping
# -------------------------------------------------------------------

ASIN_REGEXES = [
    re.compile(r"/dp/([A-Z0-9]{10})(?:[/?]|$)"),
    re.compile(r"/gp/product/([A-Z0-9]{10})(?:[/?]|$)"),
    re.compile(r"[?&]ASIN=([A-Z0-9]{10})(?:&|$)", re.IGNORECASE),
]


def extract_asin_from_url(url: str) -> Optional[str]:
    for rx in ASIN_REGEXES:
        m = rx.search(url)
        if m:
            return m.group(1)
    return None


def _extract_sell_through_from_card(a_tag: BeautifulSoup) -> str:
    """
    Try to find a span like:
      <span class="a-text-bold">500+ bought</span>
    or similar, near the product anchor on the Best Sellers page.
    """
    # Walk up a few ancestors and search within
    node = a_tag
    for _ in range(4):
        if node is None:
            break
        spans = node.find_all("span", class_="a-text-bold")
        for sp in spans:
            txt = sp.get_text(" ", strip=True)
            if "bought" in txt.lower():
                return txt
        node = node.parent

    # Fallback: first bold span after the anchor
    sib = a_tag.find_next("span", class_="a-text-bold")
    if sib:
        txt = sib.get_text(" ", strip=True)
        if "bought" in txt.lower():
            return txt

    return ""


def parse_top20_from_category_page(url: str, session: Optional[requests.Session] = None) -> List[Dict[str, Any]]:
    session = session or _session()
    soup, final_url = get_soup(url, session)
    anchors = soup.find_all("a", href=True)
    seen_asins = set()
    items: List[Dict[str, Any]] = []

    for a in anchors:
        href = a.get("href", "")
        asin = extract_asin_from_url(href)
        if not asin or asin in seen_asins:
            continue

        title = (a.get_text(strip=True) or "").strip()
        if not title:
            img = a.find("img", alt=True)
            if img and img.get("alt"):
                title = img["alt"].strip()
        if not title and a.get("title"):
            title = a["title"].strip()

        item_url = urljoin(final_url, href) if href.startswith("/") else href
        seen_asins.add(asin)

        sell_through = _extract_sell_through_from_card(a)

        items.append({
            "ASIN": asin,
            "Title": title or "",
            "URL": item_url,
            "SellThrough": sell_through,
        })
        if len(items) >= 20:
            break

    return items


def extract_price_from_soup_amzn(soup: BeautifulSoup) -> str:
    candidate_ids = [
        "priceblock_ourprice", "priceblock_dealprice", "priceblock_saleprice",
        "priceblock_pospromoprice", "sns-base-price", "corePrice_feature_div",
        "apex_desktop", "priceToPay"
    ]
    for cid in candidate_ids:
        el = soup.find(id=cid)
        if el:
            off = el.find("span", class_="a-offscreen")
            if off and off.get_text(strip=True):
                return off.get_text(strip=True)
            txt = el.get_text(" ", strip=True)
            m = re.search(r"\$\s*\d[\d,]*(?:\.\d{2})?", txt or "")
            if m:
                return m.group(0).replace(" ", "")

    for off in soup.select("span.a-offscreen"):
        val = off.get_text(strip=True)
        if re.match(r"^\$\s*\d", val):
            return val.replace(" ", "")

    meta = soup.find("meta", attrs={"itemprop": "price"})
    if meta and meta.get("content"):
        c = meta.get("content")
        return c if c.startswith("$") else f"${c}"

    og = soup.find("meta", attrs={"property": "og:price:amount"})
    if og and og.get("content"):
        c = og.get("content")
        return c if c.startswith("$") else f"${c}"

    return ""


def extract_image_from_soup_amzn(soup: BeautifulSoup) -> str:
    og = soup.find("meta", attrs={"property": "og:image"})
    if og and og.get("content"):
        return og["content"]

    link_img = soup.find("link", rel=lambda v: v and "image_src" in v)
    if link_img and link_img.get("href"):
        return link_img.get("href")

    landing = soup.find("img", id="landingImage")
    if landing:
        for attr in ("data-old-hires", "src", "data-a-dynamic-image"):
            val = landing.get(attr)
            if val and isinstance(val, str) and val.strip():
                if attr == "data-a-dynamic-image":
                    m = re.search(r'"(https:[^"]+)"\s*:', val)
                    if m:
                        return m.group(1)
                else:
                    return val

    img = soup.select_one("#imgTagWrapperId img")
    if img:
        for attr in ("data-old-hires", "src"):
            if img.get(attr):
                return img.get(attr)

    return ""


def fetch_item_details_amzn(item_url: str,
                            session: Optional[requests.Session] = None,
                            retries: int = 2,
                            delay_range=(1.2, 2.4)) -> (str, str):
    session = session or _session()
    for attempt in range(retries + 1):
        try:
            soup, _ = get_soup(item_url, session, timeout=20)
            price = extract_price_from_soup_amzn(soup)
            image = extract_image_from_soup_amzn(soup)
            return price, image
        except Exception:
            if attempt >= retries:
                return "", ""
            time.sleep(random.uniform(*delay_range))
    return "", ""


def _download_image_bytes(url: str, max_px: int = EXCEL_IMG_MAX_PX) -> Optional[BytesIO]:
    """Fetch & resize image; return PNG bytes or None."""
    try:
        if not url or not isinstance(url, str):
            return None
        resp = requests.get(url, headers={"User-Agent": random.choice(USER_AGENTS)}, timeout=15)
        resp.raise_for_status()
        img = PILImage.open(BytesIO(resp.content)).convert("RGBA")
        img.thumbnail((max_px, max_px), PILImage.LANCZOS)
        buff = BytesIO()
        img.save(buff, format="PNG")
        buff.seek(0)
        return buff
    except Exception:
        return None


# -------------------------------------------------------------------
# Persistence (saved searches as CSV + meta)
# -------------------------------------------------------------------

def _run_path(run_id: str) -> str:
    return os.path.join(SAVED_DIR, run_id)


def _meta_path(run_id: str) -> str:
    return os.path.join(_run_path(run_id), "meta.json")


def _data_path(run_id: str) -> str:
    return os.path.join(_run_path(run_id), "data.csv")


def list_saved_runs() -> List[Dict[str, Any]]:
    runs = []
    for rid in os.listdir(SAVED_DIR):
        rp = _run_path(rid)
        mp = _meta_path(rid)
        dp = _data_path(rid)
        if os.path.isdir(rp) and os.path.exists(mp) and os.path.exists(dp):
            try:
                with open(mp, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                runs.append({"id": rid, **meta})
            except Exception:
                continue
    runs.sort(key=lambda r: r.get("updated_at", ""), reverse=True)
    return runs


def save_run(run_id: str, df: pd.DataFrame, meta: Dict[str, Any]) -> None:
    os.makedirs(_run_path(run_id), exist_ok=True)
    df.to_csv(_data_path(run_id), index=False, encoding="utf-8")
    meta["updated_at"] = utc_now_str()
    with open(_meta_path(run_id), "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)


def load_run(run_id: str) -> (pd.DataFrame, Dict[str, Any]):
    df = pd.read_csv(_data_path(run_id), dtype=str).fillna("")
    with open(_meta_path(run_id), "r", encoding="utf-8") as f:
        meta = json.load(f)
    if "Rank" in df.columns:
        with pd.option_context("mode.chained_assignment", None):
            df["Rank"] = pd.to_numeric(df["Rank"], errors="coerce").fillna(0).astype(int)
    return df, meta


def new_run_meta(category_desc: str, amazon_url: str, fetched_at: str, name: Optional[str] = None) -> Dict[str, Any]:
    now = utc_now_str()
    return {
        "name": name or (category_desc or "Top 20"),
        "category_desc": category_desc or "",
        "amazon_url": amazon_url or "",
        "fetched_at": fetched_at or now,
        "created_at": now,
        "updated_at": now
    }


# -------------------------------------------------------------------
# Session state init
# -------------------------------------------------------------------

if "results" not in st.session_state:
    st.session_state.results: Optional[pd.DataFrame] = None
if "current_run_id" not in st.session_state:
    st.session_state.current_run_id = None
if "current_meta" not in st.session_state:
    st.session_state.current_meta = None
if "autosave" not in st.session_state:
    st.session_state.autosave = True
if "view_mode" not in st.session_state:
    st.session_state.view_mode = "List"
if "delay_min" not in st.session_state:
    st.session_state.delay_min = 1.0
if "delay_max" not in st.session_state:
    st.session_state.delay_max = 2.0


def maybe_autosave():
    if (
        st.session_state.autosave
        and st.session_state.current_run_id
        and st.session_state.results is not None
    ):
        save_run(st.session_state.current_run_id, st.session_state.results, st.session_state.current_meta)


# -------------------------------------------------------------------
# Sidebar: Saved searches + settings
# -------------------------------------------------------------------

with st.sidebar:
    st.header("Saved Searches")

    saved = list_saved_runs()
    selected_run = None
    if saved:
        labels = [f"{r['name']}  —  {r.get('updated_at','')}" for r in saved]
        idx = st.selectbox("Saved runs", options=list(range(len(saved))),
                           format_func=lambda i: labels[i])
        selected_run = saved[idx]

        col_s1, col_s2, col_s3 = st.columns(3)
        with col_s1:
            if st.button("Load"):
                df, meta = load_run(selected_run["id"])
                st.session_state.results = df
                st.session_state.current_run_id = selected_run["id"]
                st.session_state.current_meta = meta
                st.success(f"Loaded: {meta.get('name')}")
        with col_s2:
            if st.button("Save") and st.session_state.results is not None:
                if st.session_state.current_run_id:
                    save_run(st.session_state.current_run_id, st.session_state.results, st.session_state.current_meta)
                    st.success("Saved.")
        with col_s3:
            if st.button("Delete"):
                import shutil
                shutil.rmtree(_run_path(selected_run["id"]), ignore_errors=True)
                st.warning("Deleted. Refreshing list…")
                st.rerun()
    else:
        st.caption("No saved searches yet.")

    st.markdown("---")
    st.subheader("⚙️ Settings")
    st.session_state.autosave = st.checkbox("Autosave edits", value=st.session_state.autosave)
    st.session_state.delay_min = st.slider("Min per-item delay (sec)", 0.5, 5.0, st.session_state.delay_min, 0.1)
    st.session_state.delay_max = st.slider("Max per-item delay (sec)", 0.6, 6.0, st.session_state.delay_max, 0.1)
    if st.session_state.delay_max < st.session_state.delay_min:
        st.session_state.delay_max = st.session_state.delay_min


# -------------------------------------------------------------------
# Main layout: top bar, inputs, results
# -------------------------------------------------------------------

st.title("Amazon Top-20 Scraper → Excel (Top20 + Data)")

# Inputs row sits *below* the control bar
if "amz_url" not in st.session_state:
    st.session_state.amz_url = ""
if "category_desc" not in st.session_state:
    st.session_state.category_desc = ""

# Top control bar
bar = st.container()
with bar:
    col_b1, col_b2, col_b3 = st.columns([1, 1, 2])
    with col_b1:
        fetch_btn = st.button("Fetch Top 20", type="primary")
    with col_b2:
        export_placeholder = st.empty()
    with col_b3:
        st.session_state.view_mode = st.radio(
            "View",
            options=["List", "Grid", "Amazon-like"],
            horizontal=True,
            label_visibility="visible"
        )

# URL + name inputs
col_in1, col_in2 = st.columns([3, 2])
with col_in1:
    st.session_state.amz_url = st.text_input(
        "Amazon Best Sellers URL",
        value=st.session_state.amz_url,
        placeholder="https://www.amazon.com/gp/bestsellers/pc/....",
    )
with col_in2:
    st.session_state.category_desc = st.text_input(
        "Short Description (for your reference)",
        value=st.session_state.category_desc,
        placeholder="e.g., Single Board Computers",
    )

amz_url = st.session_state.amz_url.strip()
category_desc = st.session_state.category_desc.strip()


# -------------------------------------------------------------------
# Fetch handler
# -------------------------------------------------------------------

if fetch_btn:
    if not amz_url or not amz_url.lower().startswith("http"):
        st.error("Please enter a valid Amazon Best Sellers URL (must start with http/https).")
    else:
        session = _session()
        st.info("Fetching top 20 items from Amazon...")
        try:
            items = parse_top20_from_category_page(amz_url, session)
        except Exception as e:
            st.error(f"Failed to fetch category page: {e}")
            items = []

        out_rows = []
        ts = utc_now_str()
        total = len(items)
        progress = st.progress(0) if total else None
        status = st.empty()

        for idx, it in enumerate(items):
            rank = idx + 1
            status.write(f"Scraping {rank} of {total}…")
            price, image_url = fetch_item_details_amzn(it.get("URL", ""), session)
            img_bytes = None
            if image_url:
                img_bytes = _download_image_bytes(image_url, max_px=EXCEL_IMG_MAX_PX)

            out_rows.append({
                "Rank": rank,
                "ASIN": it.get("ASIN", ""),
                "AmazonURL": it.get("URL", ""),
                "ImageURL": image_url or "",
                "Title": it.get("Title", ""),
                "AmazonPrice": price or "",
                "SellThrough": it.get("SellThrough", ""),
                # Micro Center-related columns (blank / manual)
                "MCSKU": "",
                "MCTitle": "",
                "MCRetail": "",
                "MCCost": "",
                "Avg1_4": "",
                "AttrMatch": "",
                "Notes": "",
                # Meta
                "FetchedAt": ts,
                "CategoryDesc": category_desc,
                "AmazonBestURL": amz_url,
                # For UI images
                "ImageBytes": img_bytes.getvalue() if img_bytes else b"",
            })

            if progress:
                progress.progress(int(((idx + 1) / max(total, 1)) * 100))
            time.sleep(random.uniform(st.session_state.delay_min, st.session_state.delay_max))

        if progress:
            progress.empty()
        status.success("Top 20 fetched!")

        df = pd.DataFrame(out_rows)

        st.session_state.results = df
        st.session_state.current_run_id = None
        st.session_state.current_meta = None


# -------------------------------------------------------------------
# Excel builder (Top 20 + Data), with MC columns first in Data tab
# -------------------------------------------------------------------

def build_xlsx_two_sheets(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws_top = wb.active
    ws_top.title = "Top 20"
    ws_data = wb.create_sheet("Data")

    # Data sheet column order (MC columns first)
    cols = [
        "MCSKU",        # 1
        "MCTitle",      # 2
        "MCRetail",     # 3
        "MCCost",       # 4
        "Avg1_4",       # 5
        "AttrMatch",    # 6
        "Notes",        # 7
        "Rank",         # 8
        "ASIN",         # 9
        "AmazonURL",    # 10
        "ImageURL",     # 11
        "Title",        # 12
        "AmazonPrice",  # 13
        "SellThrough",  # 14
        "FetchedAt",    # 15
        "CategoryDesc", # 16
        "AmazonBestURL" # 17
    ]

    # Header row
    for j, c in enumerate(cols, start=1):
        ws_data.cell(row=1, column=j, value=c)

    top20 = df.copy()
    if "Rank" in top20.columns:
        top20 = top20.sort_values("Rank")
    top20 = top20.head(20).reset_index(drop=True)

    for i, r in top20.iterrows():
        values = [
            r.get("MCSKU", ""),
            r.get("MCTitle", ""),
            r.get("MCRetail", ""),
            r.get("MCCost", ""),
            r.get("Avg1_4", ""),
            r.get("AttrMatch", ""),
            r.get("Notes", ""),
            int(r.get("Rank", i + 1)),
            r.get("ASIN", ""),
            r.get("AmazonURL", ""),
            r.get("ImageURL", ""),
            r.get("Title", ""),
            r.get("AmazonPrice", ""),
            r.get("SellThrough", ""),
            r.get("FetchedAt", ""),
            r.get("CategoryDesc", ""),
            r.get("AmazonBestURL", ""),
        ]
        for j, v in enumerate(values, start=1):
            ws_data.cell(row=2 + i, column=j, value=v)

    # Some reasonable widths
    data_widths = [14, 40, 12, 12, 12, 20, 30, 6, 12, 30, 30, 50, 12, 18, 20, 24, 32]
    for j, w in enumerate(data_widths, start=1):
        ws_data.column_dimensions[get_column_letter(j)].width = w
    ws_data.freeze_panes = "A2"

    # ----- Top 20 layout (5 across, compact, banded, with sell-through) -----

    ITEMS_PER_ROW = 5
    BLOCK_ROWS = 11  # image + 10 text rows
    START_ROW = 1
    START_COL = 1

    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_mid = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=LABEL_BOLD, color="FFFFFF")
    bold_dark = Font(bold=LABEL_BOLD)
    thin = Side(style="thin", color="DDDDDD")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    band1 = PatternFill("solid", fgColor=BAND_COLOR_1)
    band2 = PatternFill("solid", fgColor=BAND_COLOR_2)
    rank_fill = PatternFill("solid", fgColor=RANK_BG)

    for c in range(START_COL, START_COL + ITEMS_PER_ROW):
        ws_top.column_dimensions[get_column_letter(c)].width = EXCEL_COL_WIDTH

    def DREF(col_idx: int, data_row: int) -> str:
        return f"Data!{get_column_letter(col_idx)}{data_row}"

    for idx in range(len(top20)):
        group = idx // ITEMS_PER_ROW
        col = START_COL + (idx % ITEMS_PER_ROW)
        base = START_ROW + group * BLOCK_ROWS
        data_row = 2 + idx
        band_fill = band1 if (group % 2 == 0) else band2

        # Image
        img_url = ws_data.cell(row=data_row, column=11).value  # ImageURL
        amazon_url = ws_data.cell(row=data_row, column=10).value  # AmazonURL

        img_buf = _download_image_bytes(img_url, max_px=EXCEL_IMG_MAX_PX)
        if img_buf:
            xl_img = XLImage(img_buf)
            xl_img.anchor = f"{get_column_letter(col)}{base}"
            ws_top.add_image(xl_img)

        cell = ws_top.cell(row=base, column=col, value=" ")
        if amazon_url:
            cell.hyperlink = amazon_url
        cell.alignment = center_mid
        cell.border = box
        cell.fill = band_fill
        ws_top.row_dimensions[base].height = EXCEL_IMG_ROW_HEIGHT

        # Rank row (with Amazon orange background)
        c = ws_top.cell(row=base + 1, column=col,
                        value=f'= "Rank: #" & {DREF(8, data_row)}')
        c.alignment = left_mid
        c.font = bold
        c.border = box
        c.fill = rank_fill

        # Sell-through
        c = ws_top.cell(row=base + 2, column=col,
                        value=f'= "Sell-through: " & {DREF(14, data_row)}')
        c.alignment = left_mid
        c.font = bold_dark
        c.border = box
        c.fill = band_fill

        # Amazon Title
        c = ws_top.cell(row=base + 3, column=col,
                        value=f'= "Amazon: " & {DREF(12, data_row)}')
        c.alignment = left_wrap
        c.border = box
        c.fill = band_fill

        # Amazon price
        c = ws_top.cell(row=base + 4, column=col,
                        value=f'= "Amazon Price: " & {DREF(13, data_row)}')
        c.alignment = left_mid
        c.font = bold_dark
        c.border = box
        c.fill = band_fill

        # MC SKU
        c = ws_top.cell(row=base + 5, column=col,
                        value=f'= "MC SKU: " & {DREF(1, data_row)}')
        c.alignment = left_mid
        c.border = box
        c.fill = band_fill

        # MC Title
        c = ws_top.cell(row=base + 6, column=col,
                        value=f'= "MC Title: " & {DREF(2, data_row)}')
        c.alignment = left_wrap
        c.border = box
        c.fill = band_fill

        # MC Retail
        c = ws_top.cell(row=base + 7, column=col,
                        value=f'= "MC Retail: " & {DREF(3, data_row)}')
        c.alignment = left_mid
        c.font = bold_dark
        c.border = box
        c.fill = band_fill

        # MC Cost
        c = ws_top.cell(row=base + 8, column=col,
                        value=f'= "MC Cost: " & {DREF(4, data_row)}')
        c.alignment = left_mid
        c.border = box
        c.fill = band_fill

        # 1–4 Avg
        c = ws_top.cell(row=base + 9, column=col,
                        value=f'= "1-4 Avg: " & {DREF(5, data_row)}')
        c.alignment = left_mid
        c.border = box
        c.fill = band_fill

        # Attributes
        c = ws_top.cell(row=base + 10, column=col,
                        value=f'= "Attributes: " & {DREF(6, data_row)}')
        c.alignment = left_wrap
        c.border = box
        c.fill = band_fill

        # Notes
        c = ws_top.cell(row=base + 11, column=col,
                        value=f'= "Notes: " & {DREF(7, data_row)}')
        c.alignment = left_wrap
        c.border = box
        c.fill = band_fill

    ws_top.freeze_panes = None

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


# -------------------------------------------------------------------
# Results view (web) – unchanged behaviour, just uses SellThrough + ImageBytes
# -------------------------------------------------------------------

if st.session_state.results is not None and not st.session_state.results.empty:
    df = st.session_state.results.copy()

    # Restore meta if available
    meta = st.session_state.current_meta or {
        "name": df.iloc[0].get("CategoryDesc") or "Top 20",
        "fetched_at": df.iloc[0].get("FetchedAt"),
        "amazon_url": df.iloc[0].get("AmazonBestURL"),
    }

    header_cols = st.columns([3, 2, 2])
    with header_cols[0]:
        st.subheader(meta.get("name", "Top 20"))
    with header_cols[1]:
        st.caption(f"Pulled at: {meta.get('fetched_at', '')}")
    with header_cols[2]:
        if meta.get("amazon_url"):
            st.caption(f"[Source: Amazon Best Sellers]({meta['amazon_url']})")

    # --- View modes ---
    mode = st.session_state.view_mode

    if mode == "List":
        show_cols = ["Rank", "Title", "AmazonPrice", "SellThrough", "ASIN"]
        st.dataframe(df[show_cols].sort_values("Rank"), use_container_width=True)

    elif mode == "Grid":
        st.markdown("#### Grid view (4 × 5)")
        # 20 items max, chunked into rows of 4
        rows = []
        tmp = df.sort_values("Rank").head(20).reset_index(drop=True)
        for i in range(0, len(tmp), 4):
            rows.append(tmp.iloc[i:i+4])

        for chunk in rows:
            cols = st.columns(len(chunk))
            for col, (_, row) in zip(cols, chunk.iterrows()):
                with col:
                    img_bytes = row.get("ImageBytes", b"")
                    if isinstance(img_bytes, (bytes, bytearray)) and img_bytes:
                        st.image(img_bytes, width=120)
                    st.markdown(f"**#{int(row['Rank'])}**")
                    if row.get("Title"):
                        st.write(row["Title"])
                    if row.get("AmazonPrice"):
                        st.write(row["AmazonPrice"])
                    if row.get("SellThrough"):
                        st.caption(row["SellThrough"])
                    if row.get("AmazonURL"):
                        st.markdown(f"[Amazon Link]({row['AmazonURL']})")
            st.markdown("---")

    else:  # "Amazon-like"
        st.markdown("#### Amazon-like vertical cards")
        tmp = df.sort_values("Rank").head(20).reset_index(drop=True)
        for _, row in tmp.iterrows():
            cols = st.columns([1, 3])
            with cols[0]:
                img_bytes = row.get("ImageBytes", b"")
                if isinstance(img_bytes, (bytes, bytearray)) and img_bytes:
                    st.image(img_bytes, width=120)
            with cols[1]:
                st.markdown(f"**#{int(row['Rank'])}**")
                if row.get("Title"):
                    st.write(row["Title"])
                price_line = row.get("AmazonPrice") or ""
                if price_line:
                    st.markdown(f"**{price_line}**")
                if row.get("SellThrough"):
                    st.caption(row["SellThrough"])
                meta_line = f"ASIN: {row.get('ASIN','')}"
                st.caption(meta_line)
                if row.get("AmazonURL"):
                    st.markdown(f"[View on Amazon]({row['AmazonURL']})")
            st.markdown("---")

    maybe_autosave()

    # Export button (Excel) – using current df
    export_placeholder.download_button(
        "Download Excel",
        data=build_xlsx_two_sheets(df),
        file_name=f"{(meta.get('name') or 'Top20').replace(' ','_')}_Top20.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
else:
    st.info("Enter an Amazon Best Sellers URL and click **Fetch Top 20** to begin.")
